import difflib
import matplotlib.pyplot as plt
import os
import numpy as np
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from joblib import Parallel, delayed
from tqdm import tqdm
import pickle
import datetime
starttime = datetime.datetime.now()

def get_label_ind(sheet_head, col_indices):
    res = []
    for lab in col_indices:
        i = 0
        for _ in sheet_head:
            if lab in _.value:
                res.append(get_column_letter(i+1))
                break
            i += 1
    return res   
def is_number(string):
    pattern = r'^-?\d+(\.\d+)?$'  # 匹配整数或小数形式的数字
    match = re.match(pattern, string)
    return match is not None
def search_company(name, ls):
    length = len(ls)
    similarity = np.zeros(length)
    for i, _ in enumerate(ls):
        try:
            similarity[i] = difflib.SequenceMatcher(autojunk=True, a=name, b=_).ratio()
        except Exception as e:
            idx = 1
            val = 'Error'
            return idx, val
    idx = np.argmax(similarity)
    val = ls[idx]
    return idx, val
# def my_fun(TAX, Name2, *args):
#     NAME_COMP = args[0]
#     col_ind_dest = args[1]
#     sheet_dest = args[2]
#     for j, _ in enumerate(TAX):
#         for i, value in enumerate(_, start=3):
#             val_name = Name2[j]
#             idx = search_company(val_name[i], NAME_COMP)        
#     return 

#####################Main Function#####################
def main():
    path = os.getcwd()
    dest_file = r'/home/egnail/proj/NiCheng/企业信息及税收/泥城企业基本信息表.xlsx'
    tax_folder = r'/home/egnail/proj/NiCheng/企业信息及税收/招商服务机构扶持明细（税收）/'
    company_file = r'/home/egnail/proj/NiCheng/企业信息及税收/企业清单（15-22年）.xlsx'
    wb_com = load_workbook(company_file)
    sheet_com = wb_com.active
    wb_dest = load_workbook(dest_file)
    sheet_dest = wb_dest.active
    col_indices1 = ['企业名称', '18位税号', '注册实体', '企业界定日期']
    col_indices2 = ['企业名称', '税号', '注册型/实地型', '户管界定日期/注册时间']
    col_ind_dest = get_label_ind(sheet_dest[2], col_indices2)
    COLS = []
    for i, _ in enumerate(col_indices1):
        for col in sheet_com.iter_cols(values_only=True):
            if _ == col[0]:
                COLS.append(col[1:])
                break
    LEN = len(COLS[0])
    COLS.insert(0, [i+1 for i in range(LEN)])
    # col_ind_dest.insert(0, 'A')
    # for j, _ in enumerate(COLS):
    #     for i, value in enumerate(_, start=3):
    #         cell = sheet_dest[col_ind_dest[j] + str(i)]
    #         cell.value = value
    NAME_COMP = COLS[1]
    #####################Tax#####################
    # ns = ['G', 'D' ,'F', 'D', 'E']  ##for 2021, 2019, 2018, 2022, 2020
    ns = ['F', 'D', 'E', 'G', 'D']  ##for 2018, 2019, 2020, 2021, 2022
    col_indices3 = []
    Name2, TAX, IDD, NameFound, IDX = [], [], [], [], []
    header_row = sheet_dest[2]
    for file in os.listdir(tax_folder):
        if '四季度' in file:
            abs_path = os.path.join(tax_folder, file)
            year = file.split('年')[0]
            col_indices3.append([_i.value for _i in header_row if year in _i.value][0])
            wb_tax = load_workbook(abs_path, data_only=True)
            sh_name = wb_tax.sheetnames[0]
            sheet_tax = wb_tax[sh_name]
            sub_name = []
            j = 0
            for col in sheet_tax.iter_cols(min_col=1, max_col=5, min_row=1, max_row=3, values_only=True):
                sub = tuple(filter(lambda x: x is not None, col))
                sub_name = ''.join(str(sub))
                if '企业名称' in sub_name:
                    ns_ = get_column_letter(j+1)
                    idd = col.index('企业名称')+1
                    IDD.append(idd)
                    break
                j += 1
            for col in sheet_tax.iter_cols(min_col=column_index_from_string(ns_), max_col=column_index_from_string(ns_), values_only=True):
                Name2.append(col[idd:])
            for col in sheet_tax.iter_cols(min_col=column_index_from_string(ns[int(year)-2018]), max_col=column_index_from_string(ns[int(year)-2018]), values_only=True):
                TAX.append(col[idd:])
    col_ind_dest = get_label_ind(header_row, col_indices3)
    # dir = os.path.join(path, '/proj/NiCheng/id_list.pkl')
    dir = path + '/proj/NiCheng/id_list.pkl'
    print(dir)
    if not os.path.exists(dir):
        for val_name in Name2:
            res = Parallel(n_jobs=64)(delayed(search_company)(_, NAME_COMP) for _ in tqdm(val_name))
            IDX.append([_[0] for _ in res])
            NameFound.append([_[1] for _ in res])
        # 保存嵌套列表到文件
        with open(path+'/proj/NiCheng/id_list.pkl', 'wb') as file:
            pickle.dump(IDX, file)
        with open(path+'/proj/NiCheng/name_list.pkl', 'wb') as file:
            pickle.dump(NameFound, file)
    else:
        # 重新加载保存的嵌套列表对象
        with open(path+'/proj/NiCheng/id_list.pkl', 'rb') as file:
            loaded_id_list = pickle.load(file)
        with open(path+'/proj/NiCheng/name_list.pkl', 'rb') as file:
            loaded_name_list = pickle.load(file)
        column_indices = [column_index_from_string(_) for _ in col_ind_dest]
        col_ind_dest1 = [get_column_letter(_+5) for _ in column_indices]
        for j, _ in enumerate(TAX):
            val_name = Name2[j]
            for i, value in enumerate(_):
                idx = loaded_id_list[j][i]
                cell = sheet_dest[col_ind_dest[j] + str(idx+3)]
                cell.value = value
                cell1 = sheet_dest[col_ind_dest1[j] + str(idx+3)]
                cell1.value = loaded_name_list[j][i]
                if i%1000 == 0:
                    print(f'{i}th row finished.')
            print('-----------------'f'No.{col_indices3[j]} finished.'+'------------------')
        wb_dest.save(dest_file)
    endtime = datetime.datetime.now()
    print (f"Total training time: {(endtime - starttime).seconds:d}")

if __name__ == '__main__':
    main()